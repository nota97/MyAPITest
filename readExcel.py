import os
import openpyxl
from openpyxl.styles import Font,colors

#读取项目路径
rootpath = os.path.split(os.path.realpath(__file__))[0]
path = os.path.join(rootpath, 'testdata')

class readExcel():
    def getExcelcase(self,filename,sheetname):
        cls=[]
        xpath=os.path.join(path,filename)
        file = openpyxl.load_workbook(xpath)
        sheet = file[sheetname]
        rows = sheet.max_row
        # sheet['A3']="http://192.168.1.21/justsy/rpc/msgPushByUser"
        # file.save('APItestcase.xlsx')
        for i in range(rows - 1):
            lst = []
            for j in range(sheet.max_column):
                lst.append(sheet.cell(i + 2, j + 1).value)
            cls.append(lst)
        print(cls)
        return cls

    @classmethod
    def addresultintoExcel(self,filename,sheetname,data):
        xpath = os.path.join(path, filename)
        file = openpyxl.load_workbook(xpath)
        sheet = file[sheetname]
        rows = sheet.max_row
        col = sheet.max_column
        sheet.cell(1, col+1).value = "result"
        for i in range(rows-1):
            if data[i] == "Pass":
                sheet.cell(i+2,col+1).font = Font(color="00FF00")
                sheet.cell(i+2, col+1).value = data[i]
            else:
                sheet.cell(i + 2, col + 1).value = data[i]
            # print(sheet.cell(i+2, col+1).value)
        file.save(rootpath+"/result/"+filename)
        return 0

if __name__ == '__main__':
    a = ["pass","false"]
    readExcel().getExcelcase("APItestcase1.xlsx", "Sheet1")
    # readExcel().addresultintoExcel("APItestcase.xlsx", "Sheet1", a)


